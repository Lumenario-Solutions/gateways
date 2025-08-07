"""
Transaction service for handling MPesa transaction operations and business logic.
"""

from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal

from mpesa.models import Transaction, CallbackLog
from core.exceptions import MPesaException, ValidationException
from core.utils.phone import normalize_phone_number, PhoneNumberError

import logging

logger = logging.getLogger(__name__)


class TransactionService:
    """
    Service for handling transaction operations and business logic.
    """

    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def get_transaction_by_id(self, transaction_id, client=None):
        """
        Get transaction by ID with optional client filter.

        Args:
            transaction_id (str): Transaction ID
            client: Optional client filter

        Returns:
            Transaction: Transaction instance
        """
        try:
            queryset = Transaction.objects.all()
            if client:
                queryset = queryset.filter(client=client)

            return queryset.get(transaction_id=transaction_id)

        except Transaction.DoesNotExist:
            raise ValidationException("Transaction not found")
        except Exception as e:
            logger.error(f"Error getting transaction {transaction_id}: {e}")
            raise MPesaException(f"Failed to retrieve transaction: {e}")

    def search_transactions(self, client=None, filters=None, page=1, page_size=20):
        """
        Search transactions with advanced filtering.

        Args:
            client: Optional client filter
            filters (dict): Search filters
            page (int): Page number
            page_size (int): Items per page

        Returns:
            dict: Search results with pagination
        """
        try:
            queryset = Transaction.objects.all()

            # Apply client filter
            if client:
                queryset = queryset.filter(client=client)

            # Apply filters
            if filters:
                queryset = self._apply_transaction_filters(queryset, filters)

            # Order by creation date (newest first)
            queryset = queryset.order_by('-created_at')

            # Pagination
            total_count = queryset.count()
            start = (page - 1) * page_size
            end = start + page_size
            transactions = queryset[start:end]

            return {
                'transactions': list(transactions),
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total_count': total_count,
                    'total_pages': (total_count + page_size - 1) // page_size,
                    'has_next': end < total_count,
                    'has_previous': page > 1
                }
            }

        except Exception as e:
            logger.error(f"Error searching transactions: {e}")
            raise MPesaException(f"Failed to search transactions: {e}")

    def _apply_transaction_filters(self, queryset, filters):
        """Apply filters to transaction queryset."""
        try:
            # Status filter
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'].upper())

            # Transaction type filter
            if filters.get('transaction_type'):
                queryset = queryset.filter(transaction_type=filters['transaction_type'].upper())

            # Date range filter
            if filters.get('date_from'):
                date_from = self._parse_date(filters['date_from'])
                queryset = queryset.filter(created_at__gte=date_from)

            if filters.get('date_to'):
                date_to = self._parse_date(filters['date_to'])
                queryset = queryset.filter(created_at__lte=date_to)

            # Amount range filter
            if filters.get('amount_min'):
                queryset = queryset.filter(amount__gte=Decimal(str(filters['amount_min'])))

            if filters.get('amount_max'):
                queryset = queryset.filter(amount__lte=Decimal(str(filters['amount_max'])))

            # Phone number filter
            if filters.get('phone_number'):
                try:
                    normalized_phone = normalize_phone_number(filters['phone_number'])
                    queryset = queryset.filter(phone_number=normalized_phone)
                except PhoneNumberError:
                    # If phone number is invalid, return empty result
                    queryset = queryset.none()

            # Reference filter
            if filters.get('reference'):
                queryset = queryset.filter(
                    Q(reference__icontains=filters['reference']) |
                    Q(mpesa_receipt_number__icontains=filters['reference'])
                )

            # Description filter
            if filters.get('description'):
                queryset = queryset.filter(description__icontains=filters['description'])

            return queryset

        except Exception as e:
            logger.error(f"Error applying transaction filters: {e}")
            raise ValidationException(f"Invalid filter parameters: {e}")

    def _parse_date(self, date_string):
        """Parse date string to datetime object."""
        try:
            # Try different date formats
            formats = [
                '%Y-%m-%d',
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%dT%H:%M:%S',
                '%Y-%m-%dT%H:%M:%SZ',
                '%Y-%m-%dT%H:%M:%S.%fZ'
            ]

            for fmt in formats:
                try:
                    return datetime.strptime(date_string, fmt)
                except ValueError:
                    continue

            raise ValueError(f"Invalid date format: {date_string}")

        except Exception as e:
            raise ValidationException(f"Invalid date format: {e}")

    def get_transaction_statistics(self, client=None, period_days=30):
        """
        Get transaction statistics for a client or overall.

        Args:
            client: Optional client filter
            period_days (int): Period in days for statistics

        Returns:
            dict: Transaction statistics
        """
        try:
            # Calculate date range
            end_date = timezone.now()
            start_date = end_date - timedelta(days=period_days)

            queryset = Transaction.objects.filter(
                created_at__gte=start_date,
                created_at__lte=end_date
            )

            if client:
                queryset = queryset.filter(client=client)

            # Basic counts
            total_count = queryset.count()
            successful_count = queryset.filter(status='SUCCESSFUL').count()
            failed_count = queryset.filter(status='FAILED').count()
            pending_count = queryset.filter(status__in=['PENDING', 'PROCESSING']).count()

            # Amount statistics
            successful_transactions = queryset.filter(status='SUCCESSFUL')
            total_amount = successful_transactions.aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')

            # Average transaction value
            avg_amount = (total_amount / successful_count) if successful_count > 0 else Decimal('0')

            # Success rate
            success_rate = (successful_count / total_count * 100) if total_count > 0 else 0

            # Transaction type breakdown
            type_breakdown = queryset.values('transaction_type').annotate(
                count=Count('transaction_id')
            ).order_by('-count')

            # Daily breakdown
            daily_stats = self._get_daily_statistics(queryset, start_date, end_date)

            return {
                'period': {
                    'start_date': start_date.isoformat(),
                    'end_date': end_date.isoformat(),
                    'days': period_days
                },
                'totals': {
                    'total_transactions': total_count,
                    'successful_transactions': successful_count,
                    'failed_transactions': failed_count,
                    'pending_transactions': pending_count,
                    'total_amount': float(total_amount),
                    'average_amount': float(avg_amount),
                    'success_rate': round(success_rate, 2)
                },
                'breakdown': {
                    'by_type': list(type_breakdown),
                    'daily': daily_stats
                }
            }

        except Exception as e:
            logger.error(f"Error getting transaction statistics: {e}")
            raise MPesaException(f"Failed to get transaction statistics: {e}")

    def _get_daily_statistics(self, queryset, start_date, end_date):
        """Get daily transaction statistics."""
        try:
            daily_stats = []
            current_date = start_date.date()
            end_date = end_date.date()

            while current_date <= end_date:
                day_start = timezone.make_aware(
                    datetime.combine(current_date, datetime.min.time())
                )
                day_end = timezone.make_aware(
                    datetime.combine(current_date, datetime.max.time())
                )

                day_transactions = queryset.filter(
                    created_at__gte=day_start,
                    created_at__lte=day_end
                )

                total_count = day_transactions.count()
                successful_count = day_transactions.filter(status='SUCCESSFUL').count()
                total_amount = day_transactions.filter(status='SUCCESSFUL').aggregate(
                    total=Sum('amount')
                )['total'] or Decimal('0')

                daily_stats.append({
                    'date': current_date.isoformat(),
                    'total_transactions': total_count,
                    'successful_transactions': successful_count,
                    'total_amount': float(total_amount)
                })

                current_date += timedelta(days=1)

            return daily_stats

        except Exception as e:
            logger.error(f"Error getting daily statistics: {e}")
            return []

    def validate_duplicate_transaction(self, client, phone_number, amount, reference=None):
        """
        Check for duplicate transactions to prevent double charging.

        Args:
            client: Client instance
            phone_number (str): Customer phone number
            amount (Decimal): Transaction amount
            reference (str): Optional reference

        Returns:
            dict: Validation result
        """
        try:
            # Check for recent transactions with same parameters
            recent_threshold = timezone.now() - timedelta(minutes=5)

            duplicate_query = Q(
                client=client,
                phone_number=normalize_phone_number(phone_number),
                amount=amount,
                created_at__gte=recent_threshold,
                status__in=['PENDING', 'PROCESSING', 'SUCCESSFUL']
            )

            if reference:
                duplicate_query &= Q(reference=reference)

            existing_transaction = Transaction.objects.filter(duplicate_query).first()

            if existing_transaction:
                return {
                    'is_duplicate': True,
                    'existing_transaction_id': str(existing_transaction.transaction_id),
                    'message': 'Duplicate transaction detected'
                }

            return {
                'is_duplicate': False,
                'message': 'No duplicate found'
            }

        except Exception as e:
            logger.error(f"Error validating duplicate transaction: {e}")
            return {
                'is_duplicate': False,
                'message': 'Validation failed, proceeding'
            }

    def reconcile_transactions(self, client=None, date_from=None, date_to=None):
        """
        Reconcile transactions and identify discrepancies.

        Args:
            client: Optional client filter
            date_from (datetime): Start date for reconciliation
            date_to (datetime): End date for reconciliation

        Returns:
            dict: Reconciliation report
        """
        try:
            queryset = Transaction.objects.all()

            if client:
                queryset = queryset.filter(client=client)

            if date_from:
                queryset = queryset.filter(created_at__gte=date_from)

            if date_to:
                queryset = queryset.filter(created_at__lte=date_to)

            # Find transactions with issues
            issues = []

            # 1. Transactions without callbacks after timeout
            timeout_threshold = timezone.now() - timedelta(minutes=10)
            pending_transactions = queryset.filter(
                status__in=['PENDING', 'PROCESSING'],
                created_at__lte=timeout_threshold
            )

            for transaction in pending_transactions:
                issues.append({
                    'transaction_id': str(transaction.transaction_id),
                    'issue_type': 'callback_timeout',
                    'description': 'Transaction pending without callback',
                    'created_at': transaction.created_at.isoformat()
                })

            # 2. Successful transactions without receipt numbers
            successful_without_receipt = queryset.filter(
                status='SUCCESSFUL',
                mpesa_receipt_number__isnull=True
            )

            for transaction in successful_without_receipt:
                issues.append({
                    'transaction_id': str(transaction.transaction_id),
                    'issue_type': 'missing_receipt',
                    'description': 'Successful transaction without receipt number',
                    'created_at': transaction.created_at.isoformat()
                })

            # 3. Failed callbacks
            failed_callbacks = CallbackLog.objects.filter(
                processed_successfully=False,
                received_at__gte=date_from if date_from else timezone.now() - timedelta(days=1)
            )

            for callback_log in failed_callbacks:
                issues.append({
                    'callback_log_id': str(callback_log.log_id),
                    'issue_type': 'callback_processing_failed',
                    'description': f'Callback processing failed: {callback_log.error_message}',
                    'received_at': callback_log.received_at.isoformat()
                })

            # Summary statistics
            summary = {
                'total_transactions': queryset.count(),
                'successful_transactions': queryset.filter(status='SUCCESSFUL').count(),
                'failed_transactions': queryset.filter(status='FAILED').count(),
                'pending_transactions': queryset.filter(status__in=['PENDING', 'PROCESSING']).count(),
                'total_issues': len(issues),
                'reconciliation_date': timezone.now().isoformat()
            }

            return {
                'summary': summary,
                'issues': issues,
                'recommendations': self._generate_reconciliation_recommendations(issues)
            }

        except Exception as e:
            logger.error(f"Error reconciling transactions: {e}")
            raise MPesaException(f"Failed to reconcile transactions: {e}")

    def _generate_reconciliation_recommendations(self, issues):
        """Generate recommendations based on reconciliation issues."""
        recommendations = []

        callback_timeout_count = len([i for i in issues if i['issue_type'] == 'callback_timeout'])
        if callback_timeout_count > 0:
            recommendations.append({
                'type': 'callback_timeout',
                'message': f'{callback_timeout_count} transactions are pending without callbacks. Consider querying their status.',
                'action': 'Query STK status for pending transactions'
            })

        missing_receipt_count = len([i for i in issues if i['issue_type'] == 'missing_receipt'])
        if missing_receipt_count > 0:
            recommendations.append({
                'type': 'missing_receipt',
                'message': f'{missing_receipt_count} successful transactions are missing receipt numbers.',
                'action': 'Review transaction data and update if possible'
            })

        failed_callback_count = len([i for i in issues if i['issue_type'] == 'callback_processing_failed'])
        if failed_callback_count > 0:
            recommendations.append({
                'type': 'failed_callbacks',
                'message': f'{failed_callback_count} callbacks failed to process.',
                'action': 'Review callback logs and reprocess if necessary'
            })

        return recommendations

    def export_transactions(self, client=None, filters=None, format='csv'):
        """
        Export transactions to specified format.

        Args:
            client: Optional client filter
            filters (dict): Export filters
            format (str): Export format ('csv', 'excel', 'json')

        Returns:
            dict: Export result with data or file path
        """
        try:
            queryset = Transaction.objects.all()

            if client:
                queryset = queryset.filter(client=client)

            if filters:
                queryset = self._apply_transaction_filters(queryset, filters)

            # Order by creation date
            queryset = queryset.order_by('-created_at')

            # Limit export size for performance
            if queryset.count() > 10000:
                raise ValidationException("Export limited to 10,000 transactions. Please refine your filters.")

            transactions = list(queryset)

            if format.lower() == 'json':
                return self._export_to_json(transactions)
            elif format.lower() == 'csv':
                return self._export_to_csv(transactions)
            elif format.lower() == 'excel':
                return self._export_to_excel(transactions)
            else:
                raise ValidationException(f"Unsupported export format: {format}")

        except Exception as e:
            logger.error(f"Error exporting transactions: {e}")
            raise MPesaException(f"Failed to export transactions: {e}")

    def _export_to_json(self, transactions):
        """Export transactions to JSON format."""
        import json

        data = []
        for transaction in transactions:
            data.append({
                'transaction_id': str(transaction.transaction_id),
                'transaction_type': transaction.transaction_type,
                'phone_number': transaction.phone_number,
                'amount': float(transaction.amount),
                'description': transaction.description,
                'reference': transaction.reference,
                'status': transaction.status,
                'mpesa_receipt_number': transaction.mpesa_receipt_number,
                'transaction_date': transaction.transaction_date.isoformat() if transaction.transaction_date else None,
                'created_at': transaction.created_at.isoformat(),
                'updated_at': transaction.updated_at.isoformat()
            })

        return {
            'format': 'json',
            'data': json.dumps(data, indent=2),
            'count': len(data)
        }

    def _export_to_csv(self, transactions):
        """Export transactions to CSV format."""
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow([
            'Transaction ID', 'Type', 'Phone Number', 'Amount', 'Description',
            'Reference', 'Status', 'MPesa Receipt', 'Transaction Date',
            'Created At', 'Updated At'
        ])

        # Write data
        for transaction in transactions:
            writer.writerow([
                str(transaction.transaction_id),
                transaction.transaction_type,
                transaction.phone_number,
                float(transaction.amount),
                transaction.description,
                transaction.reference,
                transaction.status,
                transaction.mpesa_receipt_number or '',
                transaction.transaction_date.isoformat() if transaction.transaction_date else '',
                transaction.created_at.isoformat(),
                transaction.updated_at.isoformat()
            ])

        return {
            'format': 'csv',
            'data': output.getvalue(),
            'count': len(transactions)
        }

    def _export_to_excel(self, transactions):
        """Export transactions to Excel format."""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            import io

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Transactions"

            # Headers
            headers = [
                'Transaction ID', 'Type', 'Phone Number', 'Amount', 'Description',
                'Reference', 'Status', 'MPesa Receipt', 'Transaction Date',
                'Created At', 'Updated At'
            ]

            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)

            # Data
            for row, transaction in enumerate(transactions, 2):
                worksheet.cell(row=row, column=1, value=str(transaction.transaction_id))
                worksheet.cell(row=row, column=2, value=transaction.transaction_type)
                worksheet.cell(row=row, column=3, value=transaction.phone_number)
                worksheet.cell(row=row, column=4, value=float(transaction.amount))
                worksheet.cell(row=row, column=5, value=transaction.description)
                worksheet.cell(row=row, column=6, value=transaction.reference)
                worksheet.cell(row=row, column=7, value=transaction.status)
                worksheet.cell(row=row, column=8, value=transaction.mpesa_receipt_number or '')
                worksheet.cell(row=row, column=9, value=transaction.transaction_date)
                worksheet.cell(row=row, column=10, value=transaction.created_at)
                worksheet.cell(row=row, column=11, value=transaction.updated_at)

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            return {
                'format': 'excel',
                'data': output.getvalue(),
                'count': len(transactions)
            }

        except ImportError:
            raise ValidationException("openpyxl library required for Excel export")


# Service instance - use lazy initialization to avoid database access during import
# transaction_service = TransactionService()  # Removed to prevent database access during import

def get_transaction_service():
    """Get Transaction service instance (lazy initialization)."""
    global _transaction_service
    if '_transaction_service' not in globals():
        global _transaction_service
        _transaction_service = TransactionService()
    return _transaction_service
